from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import pandas as pd
from main import run


def get_dfs():
    dfs = run()
    return dfs


def push_data(df, filename_excel='xlsx/test.xlsx', sheetname='Лист1'):
    df.to_excel(filename_excel, sheet_name=sheetname, index=False)
    return df.shape[0] + 1, df.shape[1]  # return n of columns, if [0] n of rows


def update_spreadsheet(df, filename_excel='xlsx/test.xlsx', startcol=1, startrow=1, sheetname='Sheet1', columns_name=True):
    k = 1 if columns_name is True else 0
    wb = load_workbook(filename_excel)
    for ir in range(0, len(df)):
        for ic in range(0, len(df.iloc[ir])):
            if ir == 0 and columns_name:
                wb[sheetname].cell(startrow, startcol + ic).value = df.columns[ic]
            wb[sheetname].cell(startrow + ir + k, startcol + ic).value = df.iloc[ir][ic]
    wb.save(filename_excel)


def auto_dimension(filename='xlsx/test.xlsx', sheetname='Portfolio'):
    """ https://docs-python.ru/packages/modul-openpyxl/razmer-stroki-stolbtsa/ """
    K = 0.5
    font_size = 11
    # словарь с размерами столбцов
    cols_dict = {}

    wb = load_workbook(filename)
    ws = wb[sheetname]

    # проходимся по всем строкам документа
    for row in ws.rows:
        # теперь по ячейкам каждой строки
        for cell in row:
            # получаем букву текущего столбца
            letter = cell.column_letter
            # если в ячейке записаны данные
            if cell.value:
                # устанавливаем в ячейке размер шрифта
                cell.font = Font(name='Calibri', size=font_size)
                # вычисляем количество символов, записанных в ячейку
                len_cell = len(str(cell.value))
                # длинна колонки по умолчанию, если буква
                # текущего столбца отсутствует в словаре `cols_dict`
                len_cell_dict = 0
                # смотрим в словарь c длинами столбцов
                if letter in cols_dict:
                    # если в словаре есть буква текущего столбца
                    # то извлекаем соответствующую длину
                    len_cell_dict = cols_dict[letter]

                # если текущая длина данных в ячейке
                #  больше чем длинна из словаря
                if len_cell > len_cell_dict:
                    # записываем новое значение ширины этого столбца
                    cols_dict[letter] = len_cell
                    ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                    ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                    new_width_col = len_cell * font_size ** (font_size * K)
                    # применение новой ширины столбца
                    ws.column_dimensions[cell.column_letter].width = new_width_col

def main():
    dfs = get_dfs()

    last_row, last_col = push_data(dfs[1], sheetname='Portfolio')

    update_spreadsheet(dfs[0],
                       startrow=1,
                       startcol=last_col+2,
                       sheetname='Portfolio',
                       columns_name=True)

    update_spreadsheet(dfs[2],
                       startrow=last_row+2,
                       startcol=1,
                       sheetname='Portfolio',
                       columns_name=True)

    # auto_dimension()


if __name__ == '__main__':
    main()